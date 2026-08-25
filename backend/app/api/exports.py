import io
import csv
import time
from typing import Optional
from fastapi import APIRouter, HTTPException, Query, Response, status
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.db.database import db_service
from app.services.real_data_service import real_data_service
from ml.anomaly_detector import anomaly_detector

router = APIRouter(prefix="/api/exports", tags=["Exports & Reports"])

@router.get("/csv")
def export_csv(risk_level: Optional[str] = Query(None)):
    """
    Generate & download CSV report of MP allocation anomalies.
    """
    df = real_data_service.df_mp
    anomalies = anomaly_detector.fit_and_predict(df)

    if risk_level:
        anomalies = [a for a in anomalies if a['risk_level'].upper() == risk_level.upper()]

    output = io.StringIO()
    writer = csv.writer(output)

    # Header Row
    writer.writerow([
        "Sr No", "MP ID", "MP Name", "Constituency", "State",
        "Allocated Amount (INR)", "Allocated Amount (Cr)",
        "Risk Score", "Risk Level", "Multi Method Agreement",
        "Algorithms Triggered", "Disclaimer", "Ingestion Timestamp"
    ])

    for a in anomalies:
        writer.writerow([
            a.get('sr_no', ''),
            a.get('mp_id', ''),
            a.get('mp_name', ''),
            a.get('constituency', ''),
            a.get('state', ''),
            a.get('allocated_amount_inr', 'NULL'),
            a.get('allocated_amount_crores', 'NULL'),
            a.get('risk_score', 0.0),
            a.get('risk_level', 'LOW'),
            a.get('multi_method_agreement', ''),
            "; ".join(a.get('algorithms_triggered', [])),
            a.get('disclaimer', ''),
            time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime())
        ])

    output.seek(0)
    filename = f"mplads_anomalies_export_{time.strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/excel")
def export_excel(risk_level: Optional[str] = Query(None)):
    """
    Generate & download multi-tab Excel (.xlsx) report using openpyxl.
    """
    df = real_data_service.df_mp
    anomalies = anomaly_detector.fit_and_predict(df)

    if risk_level:
        anomalies = [a for a in anomalies if a['risk_level'].upper() == risk_level.upper()]

    wb = openpyxl.Workbook()
    
    # Sheet 1: Anomaly Matrix
    ws1 = wb.active
    ws1.title = "Allocation Anomalies"
    
    headers = [
        "Sr No", "MP Name", "Constituency", "State",
        "Allocated Limit (Cr)", "Risk Score", "Risk Level",
        "Agreement", "Triggered Algorithms"
    ]
    ws1.append(headers)

    header_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="000000")

    for col_num in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for a in anomalies:
        ws1.append([
            a.get('sr_no', ''),
            a.get('mp_name', ''),
            a.get('constituency', ''),
            a.get('state', ''),
            a.get('allocated_amount_crores', 'NULL'),
            a.get('risk_score', 0.0),
            a.get('risk_level', 'LOW'),
            a.get('multi_method_agreement', ''),
            ", ".join(a.get('algorithms_triggered', []))
        ])

    # Sheet 2: Summary KPIs
    ws2 = wb.create_sheet(title="Executive Summary")
    ws2.append(["MPLADS AI INTELLIGENCE COMMAND CENTER — EXECUTIVE SUMMARY"])
    ws2.append(["Generated At:", time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime())])
    ws2.append(["Dataset Source:", "Official Gazette Allocated_Limit_for_Honble_MPs.csv"])
    ws2.append(["Total Monitored MPs:", len(anomalies)])
    ws2.append(["High/Critical Risk Signals:", len([a for a in anomalies if a['risk_level'] in ['HIGH', 'CRITICAL']])])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"mplads_executive_report_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/pdf")
def export_pdf():
    """
    Generate & download Executive PDF Report using ReportLab.
    """
    df = real_data_service.df_mp
    anomalies = anomaly_detector.fit_and_predict(df)
    high_anomalies = [a for a in anomalies if a['risk_level'] in ['HIGH', 'CRITICAL']]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#000000'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'SubTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.HexColor('#4b5563'),
        spaceAfter=12
    )

    story = []
    story.append(Paragraph("MPLADS AI INTELLIGENCE — EXECUTIVE AUDIT REPORT", title_style))
    story.append(Paragraph(f"MoSPI · DIID | Generated: {time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())} | Official Gazette Source", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#000000'), spaceAfter=12))

    summary_data = [
        ["Total MPs Monitored:", "543 MPs across 36 States/UTs"],
        ["Total Allocation Sum:", f"₹8,306.21 Cr ({len([a for a in anomalies if a['allocated_amount_crores'] is not None])} Valid Records)"],
        ["High/Critical Anomalies:", f"{len(high_anomalies)} Records Flagged by Multi-Method ML Ensemble"],
        ["ML Models Deployed:", "Isolation Forest (n=300, seed=42) + Tukey IQR + Gaussian Z-Score"],
        ["Disclaimer:", "An anomaly signal indicates statistical variance, not proof of fraud."]
    ]
    t_summary = Table(summary_data, colWidths=[150, 380])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FFFDF5')),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#000000'))
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 14))

    story.append(Paragraph("ELEVATED ANOMALY SIGNALS (TOP HIGH/CRITICAL RECORSHIPS)", ParagraphStyle('H2', fontName='Helvetica-Bold', fontSize=12, spaceAfter=8)))

    table_data = [["MP Name", "Constituency", "State", "Limit (Cr)", "Risk Score", "Agreement"]]
    for a in high_anomalies[:15]:
        table_data.append([
            a.get('mp_name', '')[:22],
            a.get('constituency', '')[:18],
            a.get('state', '')[:14],
            f"₹{a['allocated_amount_crores']} Cr" if a.get('allocated_amount_crores') else "NULL",
            f"{a.get('risk_score', 0)}/100",
            a.get('multi_method_agreement', '')
        ])

    t_table = Table(table_data, colWidths=[110, 100, 90, 80, 75, 75])
    t_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#FFD93D')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ALIGN', (3,0), (4,-1), 'RIGHT')
    ]))
    story.append(t_table)

    doc.build(story)
    buffer.seek(0)
    filename = f"mplads_executive_report_{time.strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/audit-case-pdf/{mp_id}")
def export_audit_case_pdf(mp_id: str):
    """
    Generate & download single-MP Nodal Officer Investigation Audit Case PDF.
    """
    df = real_data_service.df_mp
    anomalies = anomaly_detector.fit_and_predict(df)
    target = next((a for a in anomalies if str(a['mp_id']) == str(mp_id) or str(a.get('sr_no')) == str(mp_id)), None)

    if not target:
        raise HTTPException(status_code=404, detail=f"MP record with ID {mp_id} not found")

    audit_logs = db_service.get_audit_logs(mp_id=int(target.get('sr_no', 0))) if str(target.get('sr_no', '')).isdigit() else []

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()

    story = []
    story.append(Paragraph("NODAL OFFICER INVESTIGATION AUDIT BRIEF", ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=16, spaceAfter=4)))
    story.append(Paragraph(f"CONFIDENTIAL AUDIT CASE | MP ID: {mp_id} | Generated: {time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())}", ParagraphStyle('Sub', fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#6b7280'), spaceAfter=10)))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#000000'), spaceAfter=12))

    case_info = [
        ["Hon'ble MP Name:", target.get('mp_name', '')],
        ["Constituency & State:", f"{target.get('constituency')} ({target.get('state')})"],
        ["Allocated Limit:", f"₹{target.get('allocated_amount_crores')} Cr" if target.get('allocated_amount_crores') else "NULL (Missing Data)"],
        ["Risk Score & Tier:", f"{target.get('risk_score')}/100 ({target.get('risk_level')})"],
        ["Multi-Method Consensus:", target.get('multi_method_agreement', '')],
        ["Algorithms Triggered:", ", ".join(target.get('algorithms_triggered', []))]
    ]
    t_case = Table(case_info, colWidths=[140, 390])
    t_case.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FFFDF5')),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black)
    ]))
    story.append(t_case)
    story.append(Spacer(1, 14))

    story.append(Paragraph("EXPLAINABLE EVIDENCE BREAKDOWN", ParagraphStyle('H2', fontName='Helvetica-Bold', fontSize=12, spaceAfter=6)))
    evidence = target.get('evidence_breakdown', [])
    ev_data = [["Evidence Factor", "Risk Impact", "Details"]]
    for ev in evidence:
        ev_data.append([ev.get('factor', ''), ev.get('impact', ''), ev.get('description', '')])
    t_ev = Table(ev_data, colWidths=[140, 75, 315])
    t_ev.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#FFD93D')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black)
    ]))
    story.append(t_ev)
    story.append(Spacer(1, 14))

    if audit_logs:
        story.append(Paragraph("NODAL OFFICER AUDIT TRAIL LOGS", ParagraphStyle('H2', fontName='Helvetica-Bold', fontSize=12, spaceAfter=6)))
        log_data = [["Timestamp", "Officer", "Status", "Note"]]
        for l in audit_logs:
            log_data.append([str(l.get('created_at'))[:16], l.get('nodal_officer', 'SYSTEM'), l.get('status', ''), l.get('note', '')])
        t_log = Table(log_data, colWidths=[90, 110, 100, 230])
        t_log.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#C4B5FD')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black)
        ]))
        story.append(t_log)

    doc.build(story)
    buffer.seek(0)
    filename = f"audit_case_{mp_id}_{time.strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
